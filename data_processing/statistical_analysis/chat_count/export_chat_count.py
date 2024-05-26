 # Replace with your folder path
import os
import json
from collections import defaultdict
from openpyxl import Workbook

folder_path = r"F:\Projects\PycharmProjects\gpt_store\gptsapp_io\output"   # Replace with your folder path
output_file = 'chat_count.xlsx'

# Get all JSON files in the folder
json_files = [file for file in os.listdir(folder_path) if file.endswith('.json')]

# Create a dictionary to store chat_count and gpt count
chat_count_dict = defaultdict(int)

# Iterate over each JSON file
for json_file in json_files:
    file_path = os.path.join(folder_path, json_file)
    
    # Read the JSON file
    with open(file_path, 'r', encoding='utf-8') as file:
        data_list = json.load(file)
    
    # Iterate over each item in the list
    for data in data_list:
        # Extract the chat_count
        chat_count = data.get('chat_count', 0)
        
        # Increment the gpt count for the corresponding chat_count
        chat_count_dict[chat_count] += 1

# Create a new workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active

# Write the header
sheet['A1'] = 'chat_count'
sheet['B1'] = 'gpt_count'

# Write the data to the sheet
row = 2
for chat_count, gpt_count in sorted(chat_count_dict.items()):
    sheet.cell(row=row, column=1, value=chat_count)
    sheet.cell(row=row, column=2, value=gpt_count)
    row += 1

# Save the workbook
workbook.save(output_file)

print(f"Chat count and GPT count have been saved to {output_file}.")