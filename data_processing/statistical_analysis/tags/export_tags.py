import os
import json
from collections import defaultdict
from openpyxl import Workbook

folder_path = r"F:\Projects\PycharmProjects\gpt_store\gptsapp_io\output"   # Replace with your folder path
output_file = 'tag_count.xlsx'

# Get all JSON files in the folder
json_files = [file for file in os.listdir(folder_path) if file.endswith('.json')]

# Create a dictionary to store tag and gpt count
tag_count_dict = defaultdict(int)

# Iterate over each JSON file
for json_file in json_files:
    file_path = os.path.join(folder_path, json_file)
    
    # Read the JSON file
    with open(file_path, 'r', encoding='utf-8') as file:
        data_list = json.load(file)
    
    # Iterate over each item in the list
    for data in data_list:
        # Extract the tags
        tags = data.get('tags', [])
        
        # Increment the gpt count for each tag
        for tag in tags:
            tag_count_dict[tag] += 1

# Create a new workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active

# Write the header
sheet['A1'] = 'tag'
sheet['B1'] = 'gpt_count'

# Write the data to the sheet
row = 2
for tag, gpt_count in sorted(tag_count_dict.items()):
    sheet.cell(row=row, column=1, value=tag)
    sheet.cell(row=row, column=2, value=gpt_count)
    row += 1

# Save the workbook
workbook.save(output_file)

print(f"Tag and GPT count have been saved to {output_file}.")